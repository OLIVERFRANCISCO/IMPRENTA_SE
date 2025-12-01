"""
Módulo de exportación de datos
Soporta exportación a CSV, Excel y PDF
"""
import csv
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas


def exportar_a_csv(datos, columnas, nombre_archivo):
    """
    Exporta datos a formato CSV

    Args:
        datos (list): Lista de tuplas o diccionarios con los datos
        columnas (list): Lista de nombres de columnas
        nombre_archivo (str): Ruta del archivo a crear

    Returns:
        bool: True si fue exitoso
    """
    try:
        with open(nombre_archivo, 'w', newline='', encoding='utf-8-sig') as archivo:
            writer = csv.writer(archivo)

            # Escribir encabezados
            writer.writerow(columnas)

            # Escribir datos
            for fila in datos:
                if isinstance(fila, dict):
                    writer.writerow([fila.get(col, '') for col in columnas])
                else:
                    writer.writerow(fila)

        return True
    except Exception as e:
        print(f"Error al exportar CSV: {e}")
        return False


def exportar_a_excel(datos, columnas, nombre_archivo, titulo="Reporte"):
    """
    Exporta datos a formato Excel con formato profesional

    Args:
        datos (list): Lista de tuplas o diccionarios con los datos
        columnas (list): Lista de nombres de columnas
        nombre_archivo (str): Ruta del archivo a crear
        titulo (str): Título del reporte

    Returns:
        bool: True si fue exitoso
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Estilos
        titulo_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        titulo_fill = PatternFill(start_color='1F538D', end_color='1F538D', fill_type='solid')

        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='3A7EBF', end_color='3A7EBF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título principal
        ws.merge_cells('A1:' + chr(64 + len(columnas)) + '1')
        celda_titulo = ws['A1']
        celda_titulo.value = titulo
        celda_titulo.font = titulo_font
        celda_titulo.fill = titulo_fill
        celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Fecha y hora
        ws.merge_cells('A2:' + chr(64 + len(columnas)) + '2')
        celda_fecha = ws['A2']
        celda_fecha.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        celda_fecha.alignment = Alignment(horizontal='center')
        celda_fecha.font = Font(size=10, italic=True)
        ws.row_dimensions[2].height = 20

        # Espacio
        ws.row_dimensions[3].height = 10

        # Encabezados
        fila_header = 4
        for col_num, columna in enumerate(columnas, 1):
            celda = ws.cell(row=fila_header, column=col_num)
            celda.value = columna
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = header_alignment
            celda.border = border
            ws.column_dimensions[chr(64 + col_num)].width = max(len(str(columna)) + 5, 15)

        ws.row_dimensions[fila_header].height = 25

        # Datos
        for fila_num, fila_datos in enumerate(datos, fila_header + 1):
            if isinstance(fila_datos, dict):
                valores = [fila_datos.get(col, '') for col in columnas]
            else:
                valores = fila_datos

            for col_num, valor in enumerate(valores, 1):
                celda = ws.cell(row=fila_num, column=col_num)
                celda.value = valor
                celda.border = border
                celda.alignment = Alignment(horizontal='left', vertical='center')

                # Alternar colores de filas
                if fila_num % 2 == 0:
                    celda.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        # Ajustar anchos de columna basados en contenido
        for col_num in range(1, len(columnas) + 1):
            col_letter = chr(64 + col_num)
            max_length = 0
            for row in ws[col_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 5, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Guardar
        wb.save(nombre_archivo)
        return True

    except Exception as e:
        print(f"Error al exportar Excel: {e}")
        return False


def exportar_a_pdf(datos, columnas, nombre_archivo, titulo="Reporte", orientacion='portrait'):
    """
    Exporta datos a formato PDF con diseño profesional

    Args:
        datos (list): Lista de tuplas o diccionarios con los datos
        columnas (list): Lista de nombres de columnas
        nombre_archivo (str): Ruta del archivo a crear
        titulo (str): Título del reporte
        orientacion (str): 'portrait' o 'landscape'

    Returns:
        bool: True si fue exitoso
    """
    try:
        # Configurar página
        pagesize = A4 if orientacion == 'portrait' else (A4[1], A4[0])
        doc = SimpleDocTemplate(nombre_archivo, pagesize=pagesize,
                               rightMargin=30, leftMargin=30,
                               topMargin=30, bottomMargin=18)

        # Container para los elementos
        elementos = []

        # Estilos
        estilos = getSampleStyleSheet()
        estilo_titulo = ParagraphStyle(
            'CustomTitle',
            parent=estilos['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1F538D'),
            spaceAfter=12,
            alignment=1,  # Centrado
            fontName='Helvetica-Bold'
        )

        estilo_subtitulo = ParagraphStyle(
            'CustomSubtitle',
            parent=estilos['Normal'],
            fontSize=10,
            textColor=colors.grey,
            spaceAfter=20,
            alignment=1,
            fontName='Helvetica-Oblique'
        )

        # Título
        elementos.append(Paragraph(titulo, estilo_titulo))

        # Fecha
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        elementos.append(Paragraph(f"Generado: {fecha_actual}", estilo_subtitulo))

        elementos.append(Spacer(1, 0.2 * inch))

        # Preparar datos para la tabla
        datos_tabla = [columnas]  # Primera fila: encabezados

        for fila in datos:
            if isinstance(fila, dict):
                datos_tabla.append([str(fila.get(col, '')) for col in columnas])
            else:
                datos_tabla.append([str(val) for val in fila])

        # Crear tabla
        tabla = Table(datos_tabla)

        # Estilos de la tabla
        estilo_tabla = TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3A7EBF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),

            # Cuerpo
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1F538D')),

            # Alternar colores de filas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ])

        tabla.setStyle(estilo_tabla)
        elementos.append(tabla)

        # Generar PDF
        doc.build(elementos)
        return True

    except Exception as e:
        print(f"Error al exportar PDF: {e}")
        return False


def obtener_ruta_exportacion(tipo_archivo, nombre_base="reporte"):
    """
    Genera una ruta de archivo con timestamp

    Args:
        tipo_archivo (str): 'csv', 'xlsx' o 'pdf'
        nombre_base (str): Nombre base del archivo

    Returns:
        str: Ruta completa del archivo
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"{nombre_base}_{timestamp}.{tipo_archivo}"

    # Crear carpeta de exportaciones si no existe
    carpeta_exportaciones = Path.home() / "Documents" / "Imprenta_Reportes"
    carpeta_exportaciones.mkdir(parents=True, exist_ok=True)

    return str(carpeta_exportaciones / nombre_archivo)

